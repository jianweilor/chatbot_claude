"""
fleet_cost_suite.py

Combines the Capital Cost tool (capitalcost.py) and the Recurrent Cost tool
(recurrentcost.py) into ONE application that produces a single workbook with
exactly three tabs, in this order:

    Summary -> <user-named Capital Cost tab> -> <user-named Recurrent Cost tab>

WORKFLOW
--------
1. Ask for the Capital Cost tab name, then the Recurrent Cost tab name.
2. Build a VEHICLE ROSTER: for each vehicle, search sgcarmart_ev_combined.csv
   (same fuzzy-match search used by both original tools) and record whether
   it is MARKED or UNMARKED:
       Marked   -> no road tax, no COE (Less COE Price = $0, COE Price
                   (24mth-peak+20%) = $0, Recurrent road tax = $0)
       Unmarked -> has both (Less COE Price = latest single COE bidding
                   round premium for the vehicle's OWN COE category; COE
                   Price = that category's 24-month peak x1.2; Recurrent
                   road tax = the CSV's own road tax figure)
   Vehicle Category (Sedan/SUV/...) and COE Category (A-E) are read
   straight from each vehicle's own CSV row - nothing is assumed to be
   "Sedan" or "CAT B". This lets one Capital Cost tab hold a genuine mix
   of categories in the same run.
3. Run Capital Cost calculations for the whole roster (writes the wide
   "Car Price Estimation" table + one A-D cost-breakdown block per vehicle
   into the Capital Cost tab - unchanged logic/formulas from capitalcost.py).
4. Run Recurrent Cost calculations for the SAME roster, in the SAME order
   (writes one 24-row block per vehicle into the Recurrent Cost tab -
   unchanged logic/formulas from recurrentcost.py). Using the same roster
   for both tabs is what guarantees Summary row N always lines up with the
   same vehicle in both tabs, with no possibility of the two tabs drifting
   out of sync.
5. Generate the Summary tab using Summary.xlsx as the master template
   (its exact fonts/fills/borders/currency formats are captured and
   re-applied), with one row per roster vehicle, live-formula-linked to
   the Capital/Recurrent tabs (by their user-chosen names) - so editing a
   number in either tab updates the Summary automatically. Quantity (G) is
   left blank for the user to fill in Excel (per your instruction).
6. Reorders sheets (Summary, Capital tab, Recurrent tab) and saves as a
   new .xlsx file.

Only as many wide-table columns / A-D blocks / recurrent blocks as there
are actual vehicles are ever created - nothing is pre-allocated, and
nothing is left as an empty styled placeholder (matches the "only add
columns where there are models" requirement).

Requires:
    pip install pandas openpyxl requests beautifulsoup4
"""

from __future__ import annotations

import copy
import json
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ==========================================================================
# CONFIG
# ==========================================================================

CAPITAL_TEMPLATE_PATH = r"C:\Users\sirho\OneDrive\Desktop\SgCarMart\TCO_calculator\Capital Cost.xlsx"          # source of Capital Cost formatting
SUMMARY_TEMPLATE_PATH = "Summary.xlsx"                # source of Summary formatting
CSV_PATH = "sgcarmart_ev_combined.csv"
COE_CSV_PATH = r"C:\Users\sirho\OneDrive\Desktop\SgCarMart\TCO_calculator\COEBiddingResultsPrices.csv"
OUTPUT_XLSX = "Fleet_Cost_Summary_output.xlsx"

MAX_SHEET_NAME_LEN = 31                               # Excel's hard limit
INVALID_SHEET_NAME_CHARS = set(r"[]:*?/\\")

# --- Capital Cost layout constants (unchanged from capitalcost.py) -------
CAP_FIRST_BLOCK_ROW = 4
CAP_BLOCK_ROWS = 7
CAP_BLOCK_TEMPLATE_ROW = 11          # block #2 - the clean, quirk-free template
CAP_WIDE_FIRST_COL = "H"
CAP_WIDE_ROWS = range(4, 12)         # rows 4-11 inclusive
CAP_ACCESSORIES_TOTAL_CELL = "H27"
CAP_CAR_PRICE_BUFFER = 1.3
CAP_CONTI_BUFFER = 1.05
CAP_CONTI_PCT = 0.05
CAP_COE_FORWARD_BUFFER = 1.2
CAP_COE_PEAK_WINDOW_MONTHS = 24

CAP_REMARK_TEXT = (
    "a. Vehicle Price (w/o COE) and adding a 30% buffer): \n"
    "b. Exclude VES and EEAI rebates \n"
    "c. COE (taking peak value for last 2 years and adding 20% buffer)"
)
CAP_INTEGRATION_DESC = (
    "Integration of Accessroies (Livery, lightbar, blinker lights, PA "
    "system, Siren, Mobile Radio Set integration cost, etc.) to be fitted"
)
CAP_CONTINGENCY_DESC = "Contingency Sum \n(5% of s/n 4 and 5)"
CAP_TOTAL_DESC = "Total (s/n 4 to 6)"

# --- Recurrent Cost layout/config constants (unchanged from recurrentcost.py) ---
REC_BLOCK_HEIGHT = 24
REC_FIRST_START_ROW = 3
REC_PHYSICAL_HOLDING_QTY = 1
REC_START_FY = 2028
REC_SERVICE_YEARS = 10
REC_WARRANTY_YEARS = 1
REC_ESCALATION_RATE = 0.03
REC_ESCALATION_BASE_FY = 2026
REC_PROVISION_PCT = 0.24
REC_BATTERY_MANDAY_COST = 19735.0
REC_APPLY_BATTERY_COST_EVERY_YEAR = True
REC_BATTERY_COST_YEARS = {1, 6}
REC_ANNUAL_MILEAGE_KM = 20_000
REC_INCLUDE_DRAFT_NOTE = True

# --- Summary layout constants (captured from Summary.xlsx) ---------------
SUM_HEADER_ROW = 4
SUM_FIRST_DATA_ROW = 5
SUM_COLS = ["B", "C", "D", "E", "F", "G", "H", "I", "J"]
SUM_HEADERS = ["Vehicle Model", "Vehicle Cat", "Vehicle Type (Mark/Unmarked)",
               "Unit Price", "Recurrent Cost", "Quantity", "DB Cost",
               "Recurrent Cost", "Remark"]
SUM_CURRENCY_COLS = {"E", "F", "H", "I"}
SUM_CURRENCY_FMT = '_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'
SUM_COL_WIDTHS = {"B": 38.0, "C": 12.6, "D": 14.6, "E": 12.6, "F": 16.6,
                  "G": 10.6, "H": 18.9, "I": 23.4, "J": 32.1}
SUM_FONT_NAME = "Aptos Narrow"
SUM_FONT_SIZE = 11
SUM_HEADER_FILL = "FFFFFF00"        # yellow
SUM_TOTAL_FILL = "FFD9D9D9"         # light grey (approximates the template's theme-2 fill)


def thin_border() -> Border:
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


# ==========================================================================
# Shared CSV lookup utilities (used by both Capital and Recurrent stages)
# ==========================================================================

BRAND_FIXES = {
    "Byd": "BYD", "Bmw": "BMW", "Mg": "MG", "Gac": "GAC", "Im": "IM",
    "Nio": "NIO", "Xpeng": "XPeng", "Ev": "EV", "Iev": "iEV", "Xc40": "XC40",
    "Es90": "ES90", "Cla": "CLA",
}


def slug_to_model_name(pricing_url: str) -> str:
    m = re.search(r"/info/\d+/([^/]+)/pricing", str(pricing_url))
    slug = m.group(1) if m else str(pricing_url)
    words = [w.title() if not w.isdigit() else w for w in slug.split("-")]
    words = [BRAND_FIXES.get(w, w) for w in words]
    return " ".join(words)


def parse_money(value) -> float | None:
    """'$219,888', '$261,988\\n(w/o COE)', '$22,500 (Rebate)', '-' -> float."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value)
    match = re.search(r"[\d,]+\.?\d*", text)
    if not match:
        return None
    return float(match.group().replace(",", ""))


def parse_consumption(value) -> float | None:
    if not isinstance(value, str):
        return None
    match = re.search(r"[\d.]+", value)
    return float(match.group()) if match else None


def parse_coe_category(value) -> str | None:
    """'$129,910 (Category B COE)' -> 'B'"""
    m = re.search(r"Category\s*([A-E])", str(value), re.IGNORECASE)
    return m.group(1).upper() if m else None


def load_vehicle_df(csv_path: str) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    df["BaseModel"] = df["pricing_url"].apply(slug_to_model_name)
    df["FullName"] = df["BaseModel"] + " - " + df["SubModel"].fillna("")
    return df


def _match_rows(df: pd.DataFrame, keyword: str) -> pd.DataFrame:
    return df[df["FullName"].str.contains(keyword, case=False, na=False, regex=False)].reset_index(drop=True)


def _resolve_unique_row(df: pd.DataFrame, keyword: str) -> pd.Series | None:
    matches = _match_rows(df, keyword)
    if matches.empty:
        raise ValueError(f"No rows matched '{keyword}'.")
    if len(matches) == 1:
        return matches.iloc[0]
    kw_lower = keyword.lower()
    full_exact = matches[matches["FullName"].str.lower() == kw_lower]
    if len(full_exact) == 1:
        return full_exact.iloc[0]
    model_exact = matches[matches["BaseModel"].str.lower() == kw_lower]
    if len(model_exact) == 1:
        return model_exact.iloc[0]
    return None


def _prompt_select_row(matches: pd.DataFrame) -> pd.Series:
    print(f"\n{len(matches)} rows matched:")
    for i, r in matches.iterrows():
        print(f"  [{i}] {r['FullName']}  |  Price: {r['Current price']!r}  |  "
              f"COE: {r['COE']!r}  |  Type: {r['Vehicle Type']!r}")
    while True:
        choice = input(f"Select index [0-{len(matches) - 1}]: ").strip()
        if choice.isdigit() and 0 <= int(choice) < len(matches):
            return matches.iloc[int(choice)]
        print("Invalid selection, try again.")


def select_vehicle_interactively(df: pd.DataFrame, initial_keyword: str | None = None) -> pd.Series:
    all_models = sorted(df["BaseModel"].dropna().unique())
    keyword = initial_keyword
    while True:
        if keyword is None:
            keyword = input(
                "\nEnter a vehicle model keyword to search (e.g. 'Tesla Model Y', "
                "'BYD Seal', 'Ioniq 9'), or type 'list' to see every model in the CSV: "
            ).strip()
        if keyword.lower() == "list":
            print(f"\n{len(all_models)} models found:\n")
            for m in all_models:
                print(f"  - {m}")
            keyword = None
            continue
        try:
            row = _resolve_unique_row(df, keyword)
        except ValueError as exc:
            print(f"{exc} Try again, or type 'list' to browse.")
            keyword = None
            continue
        if row is None:
            matches = _match_rows(df, keyword)
            row = _prompt_select_row(matches)
        else:
            matches = _match_rows(df, keyword)
            if len(matches) > 1:
                print(f"'{keyword}' matched exactly - using that row automatically.")
        return row


def resolve_vehicle_noninteractive(df: pd.DataFrame, keyword: str) -> pd.Series:
    row = _resolve_unique_row(df, keyword)
    if row is None:
        matches = _match_rows(df, keyword)
        candidates = "\n".join(f"  - {r['FullName']}" for _, r in matches.iterrows())
        raise ValueError(
            f"'{keyword}' is ambiguous ({len(matches)} rows matched) and batch mode "
            f"can't prompt for a choice. Make the keyword more specific. Candidates:\n{candidates}"
        )
    return row


# ==========================================================================
# Validation
# ==========================================================================

class ValidationError(ValueError):
    pass


def validate_sheet_name(name: str, taken: set[str]) -> str:
    """Required-not-blank, <=31 chars, no invalid Excel characters, and not
    a duplicate of an already-chosen name (case-insensitive)."""
    name = (name or "").strip()
    if not name:
        raise ValidationError("Tab name cannot be blank.")
    if len(name) > MAX_SHEET_NAME_LEN:
        raise ValidationError(f"Tab name must be {MAX_SHEET_NAME_LEN} characters or fewer.")
    bad_chars = INVALID_SHEET_NAME_CHARS.intersection(set(name))
    if bad_chars:
        raise ValidationError(f"Tab name cannot contain: {' '.join(sorted(bad_chars))}")
    if name.lower() in {t.lower() for t in taken}:
        raise ValidationError(f"'{name}' is already used by another tab in this workbook.")
    return name


def prompt_sheet_name(label: str, taken: set[str]) -> str:
    while True:
        raw = input(f"{label}: ").strip()
        try:
            return validate_sheet_name(raw, taken)
        except ValidationError as exc:
            print(f"[error] {exc}")


def prompt_float(label: str, hint: str = "") -> float:
    suffix = f" ({hint})" if hint else ""
    while True:
        raw = input(f"{label}{suffix}: ").strip().replace(",", "").replace("$", "")
        try:
            return float(raw)
        except ValueError:
            print("Please enter a numeric value, e.g. 11907.60")


def prompt_float_optional(label: str, hint: str = "") -> float | None:
    suffix = f" ({hint})" if hint else ""
    raw = input(f"{label}{suffix} [Enter to skip/auto]: ").strip().replace(",", "").replace("$", "")
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        print("Please enter a numeric value, or leave blank.")
        return prompt_float_optional(label, hint)


def prompt_yes_no(label: str, default: bool) -> bool:
    default_str = "Y/n" if default else "y/N"
    raw = input(f"{label} [{default_str}]: ").strip().lower()
    if not raw:
        return default
    return raw.startswith("y")


def clean_path_input(raw: str) -> str:
    """Strips whitespace and surrounding quote marks - pasted Windows paths
    (e.g. from Explorer's 'Copy as path') often come wrapped in double
    quotes, which plain .strip() leaves in place and breaks open()."""
    text = raw.strip()
    if len(text) >= 2 and text[0] == text[-1] and text[0] in ('"', "'"):
        text = text[1:-1].strip()
    return text


def prompt_path(prompt_text: str) -> str:
    return clean_path_input(input(prompt_text))


# ==========================================================================
# Vehicle roster - shared between Capital Cost and Recurrent Cost
# ==========================================================================

@dataclass
class RosterVehicle:
    """One vehicle, resolved once, used by BOTH the Capital Cost and
    Recurrent Cost stages - this is what guarantees the two tabs (and the
    Summary rows that reference them) always stay in the same order with
    the same vehicle count."""
    display_name: str
    row: pd.Series            # the matched sgcarmart_ev_combined.csv row
    marked: bool               # True = Marked (no road tax, no COE); False = Unmarked
    vehicle_cat: str           # from CSV 'Vehicle Type' (Sedan/SUV/Commercial/...)
    coe_category: str | None   # from CSV 'COE' text (A-E) - None only if unparsable


def resolve_coe_category(row: pd.Series) -> str:
    cat = parse_coe_category(row.get("COE"))
    if cat is None:
        raise ValidationError(
            f"Could not determine a COE category (A-E) for '{row['FullName']}' "
            f"from its CSV 'COE' field ({row.get('COE')!r})."
        )
    return cat


def build_vehicle_roster(vehicle_df: pd.DataFrame, initial_keyword: str | None = None) -> list[RosterVehicle]:
    roster: list[RosterVehicle] = []
    vehicle_num = 1
    while True:
        print(f"\n--- Vehicle #{vehicle_num} ---")
        row = select_vehicle_interactively(
            vehicle_df, initial_keyword if vehicle_num == 1 else None
        )
        print(f"\nSelected: {row['FullName']}  |  Vehicle Type: {row['Vehicle Type']}  |  "
              f"COE: {row['COE']}")

        display_name = input(f"Display name for this vehicle [{row['FullName']}]: ").strip() or row["FullName"]
        marked = prompt_yes_no(
            "Is this a MARKED vehicle (no road tax, no COE)? Answer No for Unmarked "
            "(pays road tax, has COE)",
            default=True,
        )
        coe_category = resolve_coe_category(row) if not marked else parse_coe_category(row.get("COE"))
        vehicle_cat = str(row.get("Vehicle Type") or "").strip() or "Unspecified"

        roster.append(RosterVehicle(
            display_name=display_name, row=row, marked=marked,
            vehicle_cat=vehicle_cat, coe_category=coe_category,
        ))
        vehicle_num += 1

        if not prompt_yes_no("\nAdd another vehicle to this run?", default=False):
            break
    return roster


def build_roster_from_batch(vehicle_df: pd.DataFrame, specs: list[dict]) -> list[RosterVehicle]:
    roster: list[RosterVehicle] = []
    for spec in specs:
        if "keyword" not in spec:
            raise ValidationError(f"Vehicle spec missing required field 'keyword': {spec}")
        row = resolve_vehicle_noninteractive(vehicle_df, spec["keyword"])
        marked = bool(spec.get("marked", True))
        display_name = spec.get("display_name") or row["FullName"]
        coe_category = resolve_coe_category(row) if not marked else parse_coe_category(row.get("COE"))
        vehicle_cat = spec.get("vehicle_cat") or str(row.get("Vehicle Type") or "").strip() or "Unspecified"
        roster.append(RosterVehicle(
            display_name=display_name, row=row, marked=marked,
            vehicle_cat=vehicle_cat, coe_category=coe_category,
        ))
    return roster


# ==========================================================================
# COE bidding data (unchanged logic from capitalcost.py)
# ==========================================================================

def load_coe_df(path: str) -> pd.DataFrame:
    df = pd.read_csv(path)
    df["period"] = pd.PeriodIndex(df["month"], freq="M")
    return df


def latest_single_round_premium(coe_df: pd.DataFrame, category: str) -> int:
    sub = coe_df[coe_df["vehicle_class"] == f"Category {category}"]
    if sub.empty:
        raise ValidationError(f"No COE bidding data found for Category {category}.")
    sub = sub.sort_values(["period", "bidding_no"])
    return int(sub.iloc[-1]["premium"])


def peak_24mo_premium(coe_df: pd.DataFrame, category: str,
                       window_months: int = CAP_COE_PEAK_WINDOW_MONTHS) -> int:
    sub = coe_df[coe_df["vehicle_class"] == f"Category {category}"]
    if sub.empty:
        raise ValidationError(f"No COE bidding data found for Category {category}.")
    latest_period = sub["period"].max()
    cutoff = latest_period - (window_months - 1)
    window = sub[sub["period"] >= cutoff]
    return int(window["premium"].max())


# ==========================================================================
# Electricity tariff (unchanged logic from recurrentcost.py)
# ==========================================================================

def fetch_electricity_rate_or_none() -> float | None:
    try:
        import electricity_tariff as et
        soup = et.get_soup(et.URL)
        tariffs = et.extract_tariffs(soup)
        price_str = et.get_electricity_tariff_excl_gst(tariffs)
        cents = et.extract_numeric_cents(price_str)
        rate = et.cents_to_dollars_per_kwh(cents)
        if rate is not None:
            print(f"[electricity_tariff.py] Electricity tariff (excl. GST): "
                  f"{price_str} -> ${rate}/kWh")
            return rate
    except Exception as exc:  # noqa: BLE001
        print(f"[warn] Could not fetch live electricity tariff ({exc}).")
    return None


def get_electricity_rate_per_kwh() -> float:
    rate = fetch_electricity_rate_or_none()
    if rate is not None:
        return rate
    while True:
        raw = input("Enter electricity tariff manually ($/kWh, excl. GST), e.g. 0.3478: ").strip()
        try:
            return float(raw)
        except ValueError:
            print("Please enter a numeric value, e.g. 0.3478")


# ==========================================================================
# CAPITAL COST engine
# ==========================================================================

@dataclass
class CapitalVehicleResult:
    roster_vehicle: RosterVehicle
    price_listed: float
    less_coe_price: float
    ves: float
    eeai: float
    col_idx: int              # wide-table column index (8 = H)
    block_index: int          # 0-based block index
    unit_price_cell: str      # e.g. "C9" - the block's "Total (s/n 4 to 6)" cell


class CapitalCostBuilder:
    """Builds the Capital Cost tab. Re-uses capitalcost.py's exact template
    (block #2 as the clean 7-row block template, column H as the wide-table
    column template, the shared Accessories block at G18:I33/H27) but no
    longer assumes a single COE category per sheet - each vehicle's Less
    COE Price / COE Price(24mth-peak) are computed from ITS OWN COE
    category (or zeroed out if Marked), so one tab can mix categories."""

    def __init__(self, ws: Worksheet, coe_df: pd.DataFrame):
        self.ws = ws
        self.coe_df = coe_df
        self.block_template = self._capture_block_template()
        self.block_row_heights = self._capture_block_row_heights()
        self.wide_col_template = self._capture_wide_col_template()
        self._reset_sample_data()
        self.next_col_idx = 8
        self.next_block_index = 0

    # -- style capture/apply -------------------------------------------------
    @staticmethod
    def _capture_style(cell) -> dict:
        return {
            "font": copy.copy(cell.font), "fill": copy.copy(cell.fill),
            "border": copy.copy(cell.border), "alignment": copy.copy(cell.alignment),
            "number_format": cell.number_format,
        }

    @staticmethod
    def _apply_style(cell, style: dict) -> None:
        cell.font = copy.copy(style["font"])
        cell.fill = copy.copy(style["fill"])
        cell.border = copy.copy(style["border"])
        cell.alignment = copy.copy(style["alignment"])
        cell.number_format = style["number_format"]

    def _capture_block_template(self) -> dict:
        template = {}
        for offset in range(CAP_BLOCK_ROWS):
            row = CAP_BLOCK_TEMPLATE_ROW + offset
            template[offset] = {col: self._capture_style(self.ws[f"{col}{row}"]) for col in "ABCD"}
        return template

    def _capture_block_row_heights(self) -> dict:
        return {offset: self.ws.row_dimensions[CAP_BLOCK_TEMPLATE_ROW + offset].height
                for offset in range(CAP_BLOCK_ROWS)}

    def _capture_wide_col_template(self) -> dict:
        return {row: self._capture_style(self.ws[f"{CAP_WIDE_FIRST_COL}{row}"]) for row in CAP_WIDE_ROWS}

    @staticmethod
    def _blank_style(cell) -> None:
        cell.font = Font()
        cell.fill = PatternFill(fill_type=None)
        cell.border = Border()
        cell.alignment = Alignment()
        cell.number_format = "General"

    def _reset_sample_data(self) -> None:
        """Removes the 8 sample models shipped in Capital_Cost.xlsx and
        blanks all styling for unused rows/columns, so the sheet shows
        nothing until a vehicle is actually appended."""
        ws = self.ws
        for merged_range in list(ws.merged_cells.ranges):
            if str(merged_range) == "A5:D5":
                ws.unmerge_cells(str(merged_range))

        for row in range(CAP_FIRST_BLOCK_ROW, 1000):
            if all(ws.cell(row=row, column=c).value in (None, "") for c in range(1, 5)):
                next_row_blank = all(ws.cell(row=row + 1, column=c).value in (None, "") for c in range(1, 5))
                if next_row_blank and row > CAP_FIRST_BLOCK_ROW + CAP_BLOCK_ROWS:
                    break
            for col in "ABCD":
                cell = ws[f"{col}{row}"]
                cell.value = None
                self._blank_style(cell)
            ws.row_dimensions[row].height = None

        for col_idx in range(8, 200):
            col_letter = get_column_letter(col_idx)
            if all(ws[f"{col_letter}{r}"].value in (None, "") for r in CAP_WIDE_ROWS):
                if col_idx > 15:
                    break
                continue
            for r in CAP_WIDE_ROWS:
                cell = ws[f"{col_letter}{r}"]
                cell.value = None
                self._blank_style(cell)

    # -- appending one vehicle ------------------------------------------------
    def append_vehicle(self, rv: RosterVehicle) -> CapitalVehicleResult:
        row = rv.row
        price_listed = parse_money(row["Current price"])
        if price_listed is None:
            raise ValidationError(
                f"'{rv.display_name}' has no usable price in the CSV "
                f"(got {row['Current price']!r} - POA/unlisted)."
            )
        ves = parse_money(row.get("VES")) or 0.0
        eeai = parse_money(row.get("EEAI")) or 0.0

        if rv.marked:
            less_coe_price = 0
            coe_peak = None
        else:
            less_coe_price = latest_single_round_premium(self.coe_df, rv.coe_category)
            coe_peak = peak_24mo_premium(self.coe_df, rv.coe_category)

        col_idx = self.next_col_idx
        block_index = self.next_block_index
        col = get_column_letter(col_idx)

        self._write_wide_column(col_idx, rv.display_name, price_listed, less_coe_price,
                                 ves, eeai, rv.marked, coe_peak)
        self._write_block(block_index, rv.display_name, col)

        r7 = CAP_FIRST_BLOCK_ROW + CAP_BLOCK_ROWS * block_index + 5
        unit_price_cell = f"C{r7}"

        self.next_col_idx += 1
        self.next_block_index += 1

        return CapitalVehicleResult(
            roster_vehicle=rv, price_listed=price_listed, less_coe_price=less_coe_price,
            ves=ves, eeai=eeai, col_idx=col_idx, block_index=block_index,
            unit_price_cell=unit_price_cell,
        )

    def _write_wide_column(self, col_idx: int, display_name: str, price_listed: float,
                            less_coe_price: float, ves: float, eeai: float,
                            marked: bool, coe_peak: int | None) -> None:
        col = get_column_letter(col_idx)
        for row, style in self.wide_col_template.items():
            self._apply_style(self.ws[f"{col}{row}"], style)
        self.ws.column_dimensions[col].width = self.ws.column_dimensions[CAP_WIDE_FIRST_COL].width

        self.ws[f"{col}4"] = display_name
        self.ws[f"{col}5"] = price_listed
        self.ws[f"{col}6"] = less_coe_price
        self.ws[f"{col}7"] = f"=({col}5-{col}6)*{CAP_CAR_PRICE_BUFFER}"
        if eeai:
            self.ws[f"{col}8"] = f"={ves:g}+{eeai:g}"
        else:
            self.ws[f"{col}8"] = ves
        # Marked (no COE) -> forward COE budget line is also $0.
        self.ws[f"{col}9"] = 0 if marked else f"={coe_peak:g}*{CAP_COE_FORWARD_BUFFER}"
        self.ws[f"{col}10"] = f"=SUM({col}7:{col}9)"
        self.ws[f"{col}11"] = f"={col}10*{CAP_CONTI_BUFFER}"

    def _write_block(self, block_index: int, display_name: str, wide_col: str) -> None:
        start_row = CAP_FIRST_BLOCK_ROW + CAP_BLOCK_ROWS * block_index
        for offset in range(CAP_BLOCK_ROWS):
            row = start_row + offset
            for col in "ABCD":
                self._apply_style(self.ws[f"{col}{row}"], self.block_template[offset][col])
            height = self.block_row_heights[offset]
            if height is not None:
                self.ws.row_dimensions[row].height = height

        header_row, name_row = start_row, start_row + 1
        r4, r5, r6, r7 = start_row + 2, start_row + 3, start_row + 4, start_row + 5

        self.ws[f"A{header_row}"], self.ws[f"B{header_row}"] = "S/N", "Description"
        self.ws[f"C{header_row}"], self.ws[f"D{header_row}"] = "Price", "Remarks"
        self.ws[f"A{name_row}"] = f"{display_name} - Marked"

        self.ws[f"A{r4}"], self.ws[f"B{r4}"] = 4, "Marked Vehicle"
        self.ws[f"C{r4}"] = f"={wide_col}7+{wide_col}8"
        self.ws[f"D{r4}"] = CAP_REMARK_TEXT

        self.ws[f"A{r5}"], self.ws[f"B{r5}"] = 5, CAP_INTEGRATION_DESC
        self.ws[f"C{r5}"] = f"={CAP_ACCESSORIES_TOTAL_CELL}"

        self.ws[f"A{r6}"], self.ws[f"B{r6}"] = 6, CAP_CONTINGENCY_DESC
        self.ws[f"C{r6}"] = f"={CAP_CONTI_PCT}*(C{r4}+C{r5})"

        self.ws[f"A{r7}"], self.ws[f"B{r7}"] = 7, CAP_TOTAL_DESC
        self.ws[f"C{r7}"] = f"=SUM(C{r4}:C{r6})"


# ==========================================================================
# RECURRENT COST engine (unchanged logic from recurrentcost.py)
# ==========================================================================

@dataclass
class YearRow:
    service_year: int
    physical_qty: int
    fy: int
    maint_qty: int
    cost_norm: float
    consumption: float
    maint_cost: float
    electricity_cost: float
    road_tax: float
    provisional_sum: float
    others: float = 0.0

    @property
    def total(self) -> float:
        return (self.maint_cost + self.electricity_cost + self.road_tax
                + self.provisional_sum + self.others)


def cost_norm_for_year(fy: int, service_year: int, maint_lt5: float, maint_5to10: float) -> float:
    base_rate = maint_lt5 if service_year <= 5 else maint_5to10
    n = fy - REC_ESCALATION_BASE_FY + 1
    return round(base_rate * (1 + REC_ESCALATION_RATE) ** n)


def build_year_rows(maint_lt5: float, maint_5to10: float, electricity_rate: float,
                     consumption_km_per_kwh: float, road_tax_annual: float,
                     road_tax_exempt: bool) -> list[YearRow]:
    electricity_cost_per_unit = round(
        REC_ANNUAL_MILEAGE_KM / consumption_km_per_kwh * electricity_rate
    )
    effective_road_tax = 0.0 if road_tax_exempt else road_tax_annual

    rows: list[YearRow] = [YearRow(
        service_year=0, physical_qty=0, fy=REC_START_FY, maint_qty=0,
        cost_norm=0, consumption=consumption_km_per_kwh,
        maint_cost=0, electricity_cost=0, road_tax=0, provisional_sum=0,
    )]

    for sy in range(1, REC_SERVICE_YEARS + 1):
        fy = REC_START_FY + sy
        cn = cost_norm_for_year(fy, sy, maint_lt5, maint_5to10)
        maint_qty = 0 if sy <= REC_WARRANTY_YEARS else REC_PHYSICAL_HOLDING_QTY
        maint_cost = cn * maint_qty
        electricity_cost = electricity_cost_per_unit * REC_PHYSICAL_HOLDING_QTY
        road_tax = effective_road_tax * REC_PHYSICAL_HOLDING_QTY
        battery_applies = REC_APPLY_BATTERY_COST_EVERY_YEAR or sy in REC_BATTERY_COST_YEARS
        battery_component = REC_BATTERY_MANDAY_COST if battery_applies else 0.0
        provisional_sum = round(
            REC_PROVISION_PCT * cn * REC_PHYSICAL_HOLDING_QTY + battery_component * REC_PHYSICAL_HOLDING_QTY
        )
        rows.append(YearRow(
            service_year=sy, physical_qty=REC_PHYSICAL_HOLDING_QTY, fy=fy,
            maint_qty=maint_qty, cost_norm=cn, consumption=consumption_km_per_kwh,
            maint_cost=maint_cost, electricity_cost=electricity_cost,
            road_tax=road_tax, provisional_sum=provisional_sum,
        ))
    return rows


@dataclass
class RecurrentVehicleResult:
    roster_vehicle: RosterVehicle
    maint_lt5: float
    maint_5to10: float
    road_tax_exempt: bool
    rows: list[YearRow]
    subtotal: float
    block_start_row: int
    subtotal_cell: str  # e.g. "N18"


class RecurrentCostBuilder:
    """Builds the Recurrent Cost tab. Styling is generated in code (as in
    the original recurrentcost.py - it never depended on an external
    template file), so no template workbook is needed here."""

    _thin_border: Border | None = None
    _header_fill: PatternFill | None = None
    _subtotal_fill: PatternFill | None = None

    def __init__(self, ws: Worksheet, electricity_rate: float, category_label: str = ""):
        self.ws = ws
        self.electricity_rate = electricity_rate
        self._init_styles()
        self.next_start_row = REC_FIRST_START_ROW

        title = f"RECURRENT COST PROJECTION FOR {category_label}" if category_label else "RECURRENT COST PROJECTION"
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
        if REC_INCLUDE_DRAFT_NOTE:
            note_cell = ws.cell(row=2, column=6, value="Notes in red to be removed in final version")
            note_cell.font = Font(color="FF0000", size=8)

        widths = {1: 8, 2: 11, 3: 11, 4: 8, 5: 12, 6: 14, 7: 13, 8: 13,
                  9: 13, 10: 13, 11: 12, 12: 15, 13: 8, 14: 13}
        for col, width in widths.items():
            ws.column_dimensions[chr(64 + col)].width = width

    @classmethod
    def _init_styles(cls) -> None:
        if cls._thin_border is not None:
            return
        side = Side(style="thin")
        cls._thin_border = Border(top=side, bottom=side, left=side, right=side)
        cls._header_fill = PatternFill("solid", fgColor="D9E1F2")
        cls._subtotal_fill = PatternFill("solid", fgColor="FFFF00")

    def append_vehicle(self, rv: RosterVehicle, maint_lt5: float, maint_5to10: float,
                        capital_cost_projection: float | None = None) -> RecurrentVehicleResult:
        row = rv.row
        consumption = parse_consumption(row.get("Energy Consumption"))
        if consumption is None:
            raise ValidationError(
                f"Energy Consumption is missing in the CSV for '{rv.display_name}'."
            )
        road_tax = parse_money(row.get("Road tax")) or 0.0
        road_tax_exempt = rv.marked  # Marked -> no road tax (per your rule)

        year_rows = build_year_rows(
            maint_lt5=maint_lt5, maint_5to10=maint_5to10,
            electricity_rate=self.electricity_rate, consumption_km_per_kwh=consumption,
            road_tax_annual=road_tax, road_tax_exempt=road_tax_exempt,
        )

        start_row = self.next_start_row
        subtotal, next_row = self._write_vehicle_block(
            start_row, year_rows, rv.display_name, maint_lt5, maint_5to10,
            road_tax_exempt, rv.display_name, capital_cost_projection,
        )
        subtotal_row = start_row + 15
        self.next_start_row = next_row

        return RecurrentVehicleResult(
            roster_vehicle=rv, maint_lt5=maint_lt5, maint_5to10=maint_5to10,
            road_tax_exempt=road_tax_exempt, rows=year_rows, subtotal=subtotal,
            block_start_row=start_row, subtotal_cell=f"N{subtotal_row}",
        )

    def _write_vehicle_block(self, start_row: int, rows: list[YearRow], htd_or_name: str,
                              maint_lt5: float, maint_5to10: float, road_tax_exempt: bool,
                              vehicle_name: str, capital_cost_projection: float | None) -> tuple[float, int]:
        ws = self.ws
        r0, r1 = start_row, start_row + 1
        r_name = start_row + 2
        r_year0 = start_row + 3
        r_years = list(range(start_row + 4, start_row + 14))
        r_total = start_row + 14
        r_subtotal = start_row + 15
        r_notes_label = start_row + 16
        r_notes_start = start_row + 18

        top_headers = {
            1: "HTD", 2: "Service years", 3: "Physical holding Qty", 4: "FY",
            5: "Units in Comprehensive maint Qty", 6: "Capital cost projection",
            7: "Cost norms", 9: "Total recurrent requirement for complete fleet size ($)",
            14: "Total ($)",
        }
        sub_headers = {
            7: "Estimated annual Comprehensive maint cost/unit ($)",
            8: "Electricity\nConsumption (km/kWh)",
            9: "Maintenance cost \n(S$)", 10: "Electricity Cost \n(S$)",
            11: "Annual Road tax ($)", 12: "Provisional sums for accident & adhoc",
            13: "Others",
        }
        for col, text in top_headers.items():
            ws.cell(row=r0, column=col, value=text)
        for col, text in sub_headers.items():
            ws.cell(row=r1, column=col, value=text)

        for col in (1, 2, 3, 4, 5, 6, 14):
            ws.merge_cells(start_row=r0, start_column=col, end_row=r1, end_column=col)
        ws.merge_cells(start_row=r0, start_column=7, end_row=r0, end_column=8)
        ws.merge_cells(start_row=r0, start_column=9, end_row=r0, end_column=13)

        for row in (r0, r1):
            for col in range(1, 15):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = self._header_fill
                cell.border = self._thin_border

        ws.cell(row=r_name, column=1, value="TP")
        ws.cell(row=r_name, column=2, value=vehicle_name)
        ws.merge_cells(start_row=r_name, start_column=1, end_row=r_years[-1], end_column=1)
        ws.merge_cells(start_row=r_name, start_column=2, end_row=r_name, end_column=14)
        ws.cell(row=r_name, column=1).font = Font(bold=True)
        ws.cell(row=r_name, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r_name, column=2).font = Font(bold=True, size=12)
        ws.cell(row=r_name, column=2).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r_name, column=2).fill = self._header_fill

        year0 = rows[0]
        ws.cell(row=r_year0, column=2, value=year0.service_year)
        ws.cell(row=r_year0, column=3, value=year0.physical_qty)
        ws.cell(row=r_year0, column=4, value=year0.fy)

        for r_idx, yr in zip(r_years, rows[1:]):
            ws.cell(row=r_idx, column=2, value=yr.service_year)
            ws.cell(row=r_idx, column=3, value=yr.physical_qty)
            ws.cell(row=r_idx, column=4, value=yr.fy)
            ws.cell(row=r_idx, column=5, value=yr.maint_qty)
            ws.cell(row=r_idx, column=7, value=yr.cost_norm)
            ws.cell(row=r_idx, column=8, value=yr.consumption)
            ws.cell(row=r_idx, column=9, value=yr.maint_cost)
            ws.cell(row=r_idx, column=10, value=yr.electricity_cost)
            ws.cell(row=r_idx, column=11, value=yr.road_tax)
            ws.cell(row=r_idx, column=12, value=yr.provisional_sum)
            ws.cell(row=r_idx, column=13, value="-" if yr.others == 0 else yr.others)
            ws.cell(row=r_idx, column=14, value=yr.total)

        if capital_cost_projection is not None:
            ws.cell(row=r_years[0], column=6, value=capital_cost_projection)

        total_rows = rows[1:]
        ws.merge_cells(start_row=r_total, start_column=2, end_row=r_total, end_column=5)
        ws.cell(row=r_total, column=2, value="Total")
        ws.cell(row=r_total, column=2).alignment = Alignment(horizontal="right")
        ws.cell(row=r_total, column=7, value=sum(r.cost_norm for r in total_rows))
        ws.cell(row=r_total, column=9, value=sum(r.maint_cost for r in total_rows))
        ws.cell(row=r_total, column=10, value=sum(r.electricity_cost for r in total_rows))
        ws.cell(row=r_total, column=11, value=sum(r.road_tax for r in total_rows))
        ws.cell(row=r_total, column=12, value=sum(r.provisional_sum for r in total_rows))

        subtotal = sum(r.total for r in rows)
        ws.merge_cells(start_row=r_subtotal, start_column=1, end_row=r_subtotal, end_column=13)
        ws.cell(row=r_subtotal, column=1, value="Subtotal")
        ws.cell(row=r_subtotal, column=1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=r_subtotal, column=14, value=subtotal)
        for col in (1, 14):
            cell = ws.cell(row=r_subtotal, column=col)
            cell.font = Font(bold=True)
            cell.fill = self._subtotal_fill

        for row in range(r_name, r_subtotal + 1):
            for col in range(1, 15):
                cell = ws.cell(row=row, column=col)
                cell.border = self._thin_border
                if row == r_name:
                    continue
                if col == 8:
                    cell.number_format = "0.0"
                elif col not in (1, 2, 13):
                    cell.number_format = "#,##0"
        for r_idx in [r_year0] + r_years + [r_total]:
            ws.cell(row=r_idx, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=r_idx, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=r_idx, column=4).alignment = Alignment(horizontal="center")

        ws.cell(row=r_notes_label, column=1, value="Notes:")
        notes = [
            ("Electricity cost", self.electricity_rate,
             "$/kWh (SP Group tariff, excl. GST); assumes prices remain stable across the period"),
            ("Provision sum (accident & adhoc)", REC_PROVISION_PCT,
             f"{int(REC_PROVISION_PCT * 100)}% of annual comprehensive maintenance cost + "
             f"traction battery/man-day provision of ${REC_BATTERY_MANDAY_COST:,.0f}/unit"),
            ("Road Tax", 0 if road_tax_exempt else "from CSV",
             "Exempt (Marked vehicle)" if road_tax_exempt else "Unmarked vehicle - from sgcarmart_ev_combined.csv"),
            ("Maint Cost (<5 yrs)", maint_lt5, "User-provided (CMA2 contract rate)"),
            ("Maint Cost (5yrs to 10yrs)", maint_5to10, "User-provided (CMA2 contract rate)"),
        ]
        for i, (label, val, note) in enumerate(notes):
            r = r_notes_start + i
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=val)
            ws.cell(row=r, column=3, value=note)

        return subtotal, start_row + REC_BLOCK_HEIGHT


# ==========================================================================
# SUMMARY engine
# ==========================================================================

class SummaryBuilder:
    """Builds the Summary tab, re-using Summary.xlsx's exact captured
    styling (font, header/total fills, currency number format, column
    widths, borders) but sized to however many vehicles are actually in
    the roster - not a fixed 26 rows, and not split into per-category
    sections. One row per roster vehicle, in the same order used for the
    Capital Cost and Recurrent Cost tabs, so row N always refers to the
    same vehicle in all three tabs."""

    def __init__(self, ws: Worksheet, capital_tab: str, recurrent_tab: str):
        self.ws = ws
        self.capital_tab = capital_tab
        self.recurrent_tab = recurrent_tab
        self._write_header()

    def _write_header(self) -> None:
        ws = self.ws
        for col, width in SUM_COL_WIDTHS.items():
            ws.column_dimensions[col].width = width
        for col_letter, text in zip(SUM_COLS, SUM_HEADERS):
            cell = ws[f"{col_letter}{SUM_HEADER_ROW}"]
            cell.value = text
            cell.font = Font(name=SUM_FONT_NAME, size=SUM_FONT_SIZE, bold=True)
            cell.fill = PatternFill("solid", fgColor=SUM_HEADER_FILL)
            cell.border = thin_border()
            if col_letter in SUM_CURRENCY_COLS:
                cell.number_format = SUM_CURRENCY_FMT

    def write_rows(self, capital_results: list[CapitalVehicleResult],
                   recurrent_results: list[RecurrentVehicleResult]) -> int:
        """Returns the row number of the Total row."""
        ws = self.ws
        n = len(capital_results)
        for i in range(n):
            row = SUM_FIRST_DATA_ROW + i
            cap = capital_results[i]
            rec = recurrent_results[i]
            rv = cap.roster_vehicle
            wide_col = get_column_letter(cap.col_idx)

            ws[f"B{row}"] = f"='{self.capital_tab}'!{wide_col}4"
            ws[f"C{row}"] = rv.vehicle_cat
            ws[f"D{row}"] = "Marked" if rv.marked else "Unmarked"
            ws[f"E{row}"] = f"='{self.capital_tab}'!{cap.unit_price_cell}"
            ws[f"F{row}"] = f"='{self.recurrent_tab}'!{rec.subtotal_cell}"
            # G (Quantity) intentionally left blank for the user to fill in Excel.
            ws[f"H{row}"] = f"=E{row}*G{row}"
            ws[f"I{row}"] = f"=F{row}*G{row}"
            # J (Remark) intentionally left blank.

            for col_letter in SUM_COLS:
                cell = ws[f"{col_letter}{row}"]
                cell.font = Font(name=SUM_FONT_NAME, size=SUM_FONT_SIZE)
                cell.border = thin_border()
                cell.alignment = Alignment(vertical="center")
                if col_letter in SUM_CURRENCY_COLS:
                    cell.number_format = SUM_CURRENCY_FMT

        total_row = SUM_FIRST_DATA_ROW + n
        ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=7)
        ws[f"B{total_row}"] = "Total Cost"
        ws[f"H{total_row}"] = f"=SUM(H{SUM_FIRST_DATA_ROW}:H{total_row - 1})"
        ws[f"I{total_row}"] = f"=SUM(I{SUM_FIRST_DATA_ROW}:I{total_row - 1})"
        for col_letter in SUM_COLS:
            cell = ws[f"{col_letter}{total_row}"]
            cell.font = Font(name=SUM_FONT_NAME, size=SUM_FONT_SIZE, bold=True)
            cell.fill = PatternFill("solid", fgColor=SUM_TOTAL_FILL)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_letter in SUM_CURRENCY_COLS:
                cell.number_format = SUM_CURRENCY_FMT
        return total_row


# ==========================================================================
# Workbook assembly
# ==========================================================================

def save_workbook_safely(wb, path: str) -> str:
    try:
        wb.save(path)
        print(f"\nSaved Excel output to: {path}")
        return path
    except PermissionError:
        pass
    stamped_path = f"{Path(path).stem}_{datetime.now():%Y%m%d_%H%M%S}{Path(path).suffix}"
    print(f"\n[warn] '{path}' is locked (likely open in Excel, or OneDrive is syncing it).")
    wb.save(stamped_path)
    print(f"Saved Excel output to: {stamped_path} instead.")
    return stamped_path


def assemble_and_save(capital_tab_name: str, recurrent_tab_name: str,
                       capital_template_path: str, csv_path: str, coe_csv_path: str,
                       roster: list[RosterVehicle],
                       recurrent_inputs: list[dict], electricity_rate: float,
                       recurrent_category_label: str, output_path: str) -> str:
    """Orchestrates: Capital Cost tab -> Recurrent Cost tab -> Summary tab
    -> sheet reordering -> save. `recurrent_inputs` is a list (same order
    as `roster`) of {'maint_lt5':..., 'maint_5to10':..., 'capital_cost_projection':...}.
    """
    if len(recurrent_inputs) != len(roster):
        raise ValidationError(
            "Recurrent Cost inputs must be provided for exactly the same vehicles, "
            "in the same order, as the Capital Cost roster."
        )

    coe_df = load_coe_df(coe_csv_path)

    # 1. Capital Cost tab - start from the template workbook (keeps its
    #    embedded image / theme), rename its sheet.
    wb = load_workbook(capital_template_path)
    capital_ws = wb[wb.sheetnames[0]]
    capital_ws.title = capital_tab_name
    capital_builder = CapitalCostBuilder(capital_ws, coe_df)
    capital_results = [capital_builder.append_vehicle(rv) for rv in roster]

    # 2. Recurrent Cost tab - fresh sheet, code-generated styling.
    recurrent_ws = wb.create_sheet(title=recurrent_tab_name)
    recurrent_builder = RecurrentCostBuilder(recurrent_ws, electricity_rate, recurrent_category_label)
    recurrent_results = [
        recurrent_builder.append_vehicle(
            rv, ri["maint_lt5"], ri["maint_5to10"], ri.get("capital_cost_projection")
        )
        for rv, ri in zip(roster, recurrent_inputs)
    ]

    # 3. Summary tab.
    summary_ws = wb.create_sheet(title="Summary")
    summary_builder = SummaryBuilder(summary_ws, capital_tab_name, recurrent_tab_name)
    summary_builder.write_rows(capital_results, recurrent_results)

    # 4. Order: Summary, Capital, Recurrent.
    wb._sheets = [summary_ws, capital_ws, recurrent_ws]
    wb.active = 0

    return save_workbook_safely(wb, output_path)


# ==========================================================================
# CLI / orchestration
# ==========================================================================

def main_interactive() -> None:
    print("=== Fleet Cost Suite (Capital + Recurrent + Summary) ===")

    capital_template_path = clean_path_input(CAPITAL_TEMPLATE_PATH)
    if not Path(capital_template_path).exists():
        capital_template_path = prompt_path(
            f"'{capital_template_path}' not found. Enter full path to the Capital Cost template workbook: "
        )
    csv_path = clean_path_input(CSV_PATH)
    if not Path(csv_path).exists():
        csv_path = prompt_path(f"'{csv_path}' not found. Enter full path to sgcarmart_ev_combined.csv: ")
    coe_csv_path = clean_path_input(COE_CSV_PATH)
    if not Path(coe_csv_path).exists():
        coe_csv_path = prompt_path(
            f"'{coe_csv_path}' not found. Enter full path to COEBiddingResultsPrices.csv: "
        )

    taken = {"Summary"}
    capital_tab_name = prompt_sheet_name("What would you like to name the Capital Cost tab?", taken)
    taken.add(capital_tab_name)
    recurrent_tab_name = prompt_sheet_name("What would you like to name the Recurrent Cost tab?", taken)

    vehicle_df = load_vehicle_df(csv_path)

    print("\n--- Building vehicle roster (used for both Capital and Recurrent Cost) ---")
    roster = build_vehicle_roster(vehicle_df)

    print("\n--- Recurrent Cost inputs (same vehicles, same order) ---")
    recurrent_category_label = input(
        "Label for the Recurrent Cost tab title, e.g. 'TP Fleet FY28' [optional]: "
    ).strip()
    electricity_rate = get_electricity_rate_per_kwh()

    recurrent_inputs = []
    for rv in roster:
        print(f"\n=== {rv.display_name} ({'Marked' if rv.marked else 'Unmarked'}) ===")
        maint_lt5 = prompt_float("Maint Cost (<5 yrs)", "e.g. 11907.60")
        maint_5to10 = prompt_float("Maint Cost (5yrs to 10yrs)", "e.g. 12141.60")
        capital_cost_projection = prompt_float_optional(
            "Capital cost projection (informational only, shown in Year 1)", "e.g. 254295"
        )
        recurrent_inputs.append({
            "maint_lt5": maint_lt5, "maint_5to10": maint_5to10,
            "capital_cost_projection": capital_cost_projection,
        })

    if not prompt_yes_no("\nGenerate Summary and save the workbook now?", default=True):
        print("Cancelled - nothing was saved.")
        return

    output_path = prompt_path(f"Output file path [{OUTPUT_XLSX}]: ") or OUTPUT_XLSX

    saved_path = assemble_and_save(
        capital_tab_name, recurrent_tab_name, capital_template_path, csv_path, coe_csv_path,
        roster, recurrent_inputs, electricity_rate, recurrent_category_label, output_path,
    )
    print(f"\nDone. Workbook tabs: Summary, {capital_tab_name}, {recurrent_tab_name}")
    print(f"Saved to: {saved_path}")


def main_batch(batch_path: str) -> None:
    with open(batch_path, "r", encoding="utf-8") as f:
        spec = json.load(f)

    capital_template_path = spec.get("capital_template_path", CAPITAL_TEMPLATE_PATH)
    csv_path = spec.get("csv_path", CSV_PATH)
    coe_csv_path = spec.get("coe_csv_path", COE_CSV_PATH)
    output_path = spec.get("output_path", OUTPUT_XLSX)

    taken = {"Summary"}
    capital_tab_name = validate_sheet_name(spec["capital_tab_name"], taken)
    taken.add(capital_tab_name)
    recurrent_tab_name = validate_sheet_name(spec["recurrent_tab_name"], taken)

    vehicle_df = load_vehicle_df(csv_path)
    vehicle_specs = spec.get("vehicles")
    if not vehicle_specs:
        raise ValidationError("Batch spec has no 'vehicles' list.")
    roster = build_roster_from_batch(vehicle_df, vehicle_specs)

    electricity_rate = spec.get("electricity_rate")
    if electricity_rate is None:
        electricity_rate = fetch_electricity_rate_or_none()
    if electricity_rate is None:
        raise ValidationError(
            "Could not fetch a live electricity tariff and none was supplied "
            "('electricity_rate' in the batch JSON)."
        )

    recurrent_inputs = []
    for vspec in vehicle_specs:
        for required in ("maint_lt5", "maint_5to10"):
            if required not in vspec:
                raise ValidationError(f"Vehicle spec missing required field '{required}': {vspec}")
        recurrent_inputs.append({
            "maint_lt5": float(vspec["maint_lt5"]),
            "maint_5to10": float(vspec["maint_5to10"]),
            "capital_cost_projection": vspec.get("capital_cost_projection"),
        })

    saved_path = assemble_and_save(
        capital_tab_name, recurrent_tab_name, capital_template_path, csv_path, coe_csv_path,
        roster, recurrent_inputs, electricity_rate, spec.get("recurrent_category_label", ""),
        output_path,
    )
    print(f"\nDone. Workbook tabs: Summary, {capital_tab_name}, {recurrent_tab_name}")
    print(f"Saved to: {saved_path}")


def main() -> None:
    args = sys.argv[1:]
    if args and args[0] == "--batch":
        if len(args) < 2:
            sys.exit("Usage: python fleet_cost_suite.py --batch <path_to_spec.json>")
        main_batch(args[1])
        return
    main_interactive()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nCancelled.")
        sys.exit(1)
    except ValidationError as exc:
        print(f"\n[error] {exc}")
        sys.exit(1)